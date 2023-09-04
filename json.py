# import pandas as pd

# # Load Excel file into DataFrame
# df = pd.read_excel('C:\Users\HP\Desktop\SHE\C++\demo1.xlsx')

# # Convert DataFrame to JSON string
# json_data = df.to_json(orient='records')

# # Save JSON data to a file
# with open('output.json', 'w') as file:
#     file.write(json_data)

import json
import openpyxl

def excel_to_json(excel_file, sheet_name):
  """Converts an Excel file to JSON.

  Args:
    excel_file: The path to the Excel file.
    sheet_name: The name of the sheet to convert.

  Returns:
    A JSON string of the Excel data.
  """

  wb = openpyxl.load_workbook(excel_file)
  sheet = wb[sheet_name]

  data = {}
  for row in range(sheet.max_row):
    row_data = {}
    for col in range(sheet.max_column):
      cell_value = sheet.cell(row, col).value
      row_data[sheet.cell(1, col).value] = cell_value
    data[row] = row_data

  json_data = json.dumps(data)

  return json_data


if __name__ == "__main__":
  excel_file = "data.xlsx"
  sheet_name = "Sheet1"

  json_data = excel_to_json(excel_file, sheet_name)

  print(json_data)

